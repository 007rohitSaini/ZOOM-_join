from selenium import webdriver
import time
import openpyxl
from datetime import datetime
import numpy as np
import pyautogui
import cv2


driver = webdriver.Chrome("chromedriver.exe")

status = "Not_in_class"

def open_teams():
    driver.get("https://teams.microsoft.com/_?culture=en-in&country=IN&lm=deeplink&lmsrc=NeutralHomePageWeb&cmpid=WebSignIn#/school//?ctx=teamsGrid")
    time.sleep(5)
    login = driver.find_element_by_name("loginfmt")
    login.clear()
    login.send_keys(teams_id)
    driver.find_element_by_id("idSIButton9").click()
    password = driver.find_element_by_name("passwd")
    password.clear()
    password.send_keys(teams_password)
    time.sleep(1)
    driver.find_element_by_id("idSIButton9").click()
    time.sleep(1)
    driver.find_element_by_id("idBtn_Back").click()
    time.sleep(1)


def day():
    day = datetime.today().strftime('%A')
    # Monday

    if day == "Monday":
        column = 4
        return column
    # Tuesday

    elif day == "Tuesday":
        column = 5
        return column
    # Wednesday

    elif day == "Wednesday":
        column = 6
        return column
    # Thusday

    elif day == "Thursday":
        column = 7
        return column

    # Friday

    elif day == "Friday":
        column = 8
        return column

    else:
        column = None
        return column


def get_class_name(column):
    wb_obj = openpyxl.load_workbook(
        timtable)
    sheet_obj = wb_obj.active
    time = datetime.now().time().strftime('%H:%M')
    hours = time.replace(":", "")
    hours = int(hours)
    classes = ['English', 'Physics', 'Chemistry', 'Maths', 'Comp']
    if hours in range(910, 950):
        cell_obj = sheet_obj.cell(row=column, column=3)
        for classes in classes:
            if classes in cell_obj.value:
                Meeting = classes
                return Meeting
    elif hours in range(1000, 1050):
        cell_obj = sheet_obj.cell(row=column, column=4)
        for classes in classes:
            if classes in cell_obj.value:
                Meeting = classes
                return Meeting
    elif hours in range(1100, 1150):
        cell_obj = sheet_obj.cell(row=column, column=5)
        for classes in classes:
            if classes in cell_obj.value:
                Meeting = classes
                return Meeting
    elif hours in range(1200, 1250):
        cell_obj = sheet_obj.cell(row=column, column=6)
        for classes in classes:
            if classes in cell_obj.value:
                Meeting = classes
                return Meeting
    elif hours in range(1300, 1350):
        cell_obj = sheet_obj.cell(row=column, column=7)
        for classes in classes:
            if classes in cell_obj.value:
                Meeting = classes
                return Meeting
    else:
        Meeting = "Go have some sleep!!"
        return Meeting


def join_class(Meet):
    if Meet == "English":
        driver.get("https://teams.microsoft.com/l/channel/19%3af218aadaeeb64658932bb5426ad1073e%40thread.tacv2/ENGLISH?groupId=a70c1404-63ae-42c9-a61c-692f45fb9822&tenantId=2067488e-4019-4ef5-82fd-6ed73c6feeb8")
       
    elif Meet == "Maths":
        driver.get("https://teams.microsoft.com/l/channel/19%3a4f58757920de41f3ab0fb3b3b4e34139%40thread.tacv2/General?groupId=5abcd413-64fd-4fe5-a925-d1a0991325c1&tenantId=2067488e-4019-4ef5-82fd-6ed73c6feeb8")
       
    elif Meet == "Physics":
        driver.get("https://teams.microsoft.com/l/channel/19%3ade148554c5534005ab85d9324a57b804%40thread.tacv2/PHYSICS?groupId=a70c1404-63ae-42c9-a61c-692f45fb9822&tenantId=2067488e-4019-4ef5-82fd-6ed73c6feeb8")
        
    elif Meet == "Chemistry":
        driver.get("https://teams.microsoft.com/l/channel/19%3abb10e6022af0459e85b9565b72bc5dbe%40thread.tacv2/CHEMISTRY?groupId=a70c1404-63ae-42c9-a61c-692f45fb9822&tenantId=2067488e-4019-4ef5-82fd-6ed73c6feeb8")
        
    elif Meet == "Comp":
        driver.get("https://teams.microsoft.com/l/channel/19%3adfbec3c115ae4d5eab324378af7862dd%40thread.tacv2/General?groupId=67c94a9c-c66c-4304-a80f-2393665969b9&tenantId=2067488e-4019-4ef5-82fd-6ed73c6feeb8")
        
    else:
        driver.close()

    driver.maximize_window()
    pyautogui.press('enter')
    driver.find_element_by_id("openTeamsClientInBrowser").click()
    
    


def click_join():
    image = pyautogui.screenshot()
    image = cv2.cvtColor(np.array(image),
                         cv2.COLOR_RGB2BGR)
    cv2.imwrite('C:\\Users\\ROHIT\\Desktop\\Class_join\\screen.png', image)
    img_rgb = cv2.imread('screen.png')
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
    template = cv2.imread("join.png", 0)
    w, h = template.shape[::-1]

    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    threshold = 0.8
    loc = np.where(res >= threshold)
    for pt in zip(*loc[::-1]):
        cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0, 0, 255), 2)

        pyautogui.moveTo(pt[0] + 10, pt[1] + 10)
        pyautogui.click()   
    cv2.imwrite('res.png', img_rgb)
    time.sleep(3)
    pyautogui.press('tab')
    pyautogui.press('enter')

    time.sleep(5)
    video = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/div[1]/div/calling-pre-join-screen/div/div/div[2]/div[1]/div[2]/div/div/section/div[2]/toggle-button[1]/div/button/span[1]')
    if video.get_attribute("title") == 'Turn camera off':
        driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/div[1]/div/calling-pre-join-screen/div/div/div[2]/div[1]/div[2]/div/div/section/div[2]/toggle-button[1]/div/button').click()
    
    time.sleep(2)
    driver.find_element_by_xpath("/html/body/div[2]/div[2]/div[2]/div[1]/div/calling-pre-join-screen/div/div/div[2]/div[1]/div[2]/div/div/section/div[1]/div/div/button").click()

if __name__ == "__main__":

    column = day()
    res = get_class_name(column)
    print(res)
    open_teams()
    time.sleep(3)
    is_there = join_class(res)
    time.sleep(5)
    
    click_join()
